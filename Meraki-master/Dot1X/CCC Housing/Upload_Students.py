# Description: This software is designed to
# automate upload students to CCC Housing Dot1X Meraki authentication for Perry Hall.
#
# Written by: Tory Davenport for First Light Fiber
# Last Updated: 8/13/2019
# 
#
from openpyxl import *
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import *
import time
from time import sleep
import sys

# Open workbook, create sheet object 
try:
	workbook_obj = load_workbook("./Student-Upload.xlsx")
	sheet_obj = workbook_obj["Student_Accounts"]
except PermissionError:
	print("Please close the file while running this script.")
except NameError:
	print("FAIL")
_min = 1
_max = 5

# Create a class to handle row data
class Data_Row:
		
	# Initialize row object
	def __init__(self, _min, _max):
		row_num = row_num
		# Use to pull individual cells for a given row
		self.last = sheet_obj.cell(row = row_num, column = 1) # First Name
		self.first = sheet_obj.cell(row = row_num, column = 2) # Last name
		self.email_address = sheet_obj.cell(row = row_num, column = 3)
		self.description = sheet_obj.cell(row = row_num, column = 4) # First + Last = Username
		self.expire_days = sheet_obj.cell(row = row_num, column = 5) # How many days this account will be authorized for
		
		# Use to print entire rows with formatting	
		self.row_formatted = f"{self.last.value} {self.first.value} {self.email_address.value} {self.description.value} {self.expire_days.value}"

# Build Stack
row1 = Data_Row(int(value))
row2 = Data_Row(int(value+1))
#row3 = Data_Row(int(value+2))
#row4 = Data_Row(int(value+3))
#row5 = Data_Row(int(value+4))
#row6 = Data_Row(int(value+5))
#row7 = Data_Row(int(value+6))
#row8 = Data_Row(int(value+7))
#row9 = Data_Row(int(value+8))
#row10 = Data_Row(int(value+10))
#row11 = Data_Row(int(value+11))
#row12 = Data_Row(int(value+12))
#row13 = Data_Row(int(value+13))
#row14 = Data_Row(int(value+14))
#row15 = Data_Row(int(value+15))
#row16 = Data_Row(int(value+16))
#row17 = Data_Row(int(value+17))
#row18 = Data_Row(int(value+18))
#row19 = Data_Row(int(value+19))
#row20 = Data_Row(int(value+20))
#row21 = Data_Row(int(value+21))
#row22 = Data_Row(int(value+22))
#row23 = Data_Row(int(value+23))
#row24 = Data_Row(int(value+24))
#row25 = Data_Row(int(value+25))
#row26 = Data_Row(int(value+26))
#row27 = Data_Row(int(value+27))
#row28 = Data_Row(int(value+28))
#row29 = Data_Row(int(value+29))
#row30 = Data_Row(int(value+30))
#row31 = Data_Row(int(value+31))
#row32 = Data_Row(int(value+32))
#row33 = Data_Row(int(value+33))
#row34 = Data_Row(int(value+34))
#row35 = Data_Row(int(value+35))
#row36 = Data_Row(int(value+36))
#row37 = Data_Row(int(value+37))
#row38 = Data_Row(int(value+38))
#row39 = Data_Row(int(value+39))
#row40 = Data_Row(int(value+40))
#row41 = Data_Row(int(value+41))
#row42 = Data_Row(int(value+42))
#row43 = Data_Row(int(value+43))
#row44 = Data_Row(int(value+44))
#row45 = Data_Row(int(value+45))
#row46 = Data_Row(int(value+46))
#row47 = Data_Row(int(value+47))
#row48 = Data_Row(int(value+48))
#row49 = Data_Row(int(value+49))
#row50 = Data_Row(int(value+50))
#row51 = Data_Row(int(value+51))
#row52 = Data_Row(int(value+52))
#row53 = Data_Row(int(value+53))
#row54 = Data_Row(int(value+54))
#row55 = Data_Row(int(value+55))
#row56 = Data_Row(int(value+56))
#row57 = Data_Row(int(value+57))
#row58 = Data_Row(int(value+58))
#row59 = Data_Row(int(value+59))
#row60 = Data_Row(int(value+60))
#row61 = Data_Row(int(value+61))
#row62 = Data_Row(int(value+62))
#row63 = Data_Row(int(value+63))
#row64 = Data_Row(int(value+64))
#row65 = Data_Row(int(value+65))
#row66 = Data_Row(int(value+66))
#row67 = Data_Row(int(value+67))
#row68 = Data_Row(int(value+68))
#row69 = Data_Row(int(value+69))
#row70 = Data_Row(int(value+70))
#row71 = Data_Row(int(value+71))
#row72 = Data_Row(int(value+72))
#row73 = Data_Row(int(value+73))
#row74 = Data_Row(int(value+74))
#row75 = Data_Row(int(value+75))
#row76 = Data_Row(int(value+76))
#row77 = Data_Row(int(value+77))
#row78 = Data_Row(int(value+78))
#row79 = Data_Row(int(value+79))
#row80 = Data_Row(int(value+80))
#row81 = Data_Row(int(value+81))
#row82 = Data_Row(int(value+82))
#row83 = Data_Row(int(value+83))
#row84 = Data_Row(int(value+84))
#row85 = Data_Row(int(value+85))
#row86 = Data_Row(int(value+86))
#row87 = Data_Row(int(value+87))
#row88 = Data_Row(int(value+88))
#row89 = Data_Row(int(value+89))
#row90 = Data_Row(int(value+90))
#row91 = Data_Row(int(value+91))
#row92 = Data_Row(int(value+92))
#row93 = Data_Row(int(value+93))
#row94 = Data_Row(int(value+94))
#row95 = Data_Row(int(value+95))
#row96 = Data_Row(int(value+96))
#row97 = Data_Row(int(value+97))
#row98 = Data_Row(int(value+98))
#row99 = Data_Row(int(value+99))
#row100 = Data_Row(int(value+100))
#row101 = Data_Row(int(value+101))
#row102 = Data_Row(int(value+102))
#row103 = Data_Row(int(value+103))
#row104 = Data_Row(int(value+104))
#row105 = Data_Row(int(value+105))
#row106 = Data_Row(int(value+106))
#row107 = Data_Row(int(value+107))
#row108 = Data_Row(int(value+108))
#row109 = Data_Row(int(value+109))
#row110 = Data_Row(int(value+110))
#row111 = Data_Row(int(value+111))
#row112 = Data_Row(int(value+112))
#row113 = Data_Row(int(value+113))
#row114 = Data_Row(int(value+114))
#row115 = Data_Row(int(value+115))
#row116 = Data_Row(int(value+116))
#row117 = Data_Row(int(value+117))
#row118 = Data_Row(int(value+118))
#row119 = Data_Row(int(value+119))
#row120 = Data_Row(int(value+120))
#row121 = Data_Row(int(value+121))
#row122 = Data_Row(int(value+122))
#row123 = Data_Row(int(value+123))
#row124 = Data_Row(int(value+124))
#row125 = Data_Row(int(value+125))
#row126 = Data_Row(int(value+126))
#row127 = Data_Row(int(value+127))
#row128 = Data_Row(int(value+128))
#row129 = Data_Row(int(value+129))
#row130 = Data_Row(int(value+130))
#row131 = Data_Row(int(value+131))
#row132 = Data_Row(int(value+132))
#row133 = Data_Row(int(value+133))
#row134 = Data_Row(int(value+134))
#row135 = Data_Row(int(value+135))
#row136 = Data_Row(int(value+136))
#row137 = Data_Row(int(value+137))
#row138 = Data_Row(int(value+138))
#row139 = Data_Row(int(value+139))
#row140 = Data_Row(int(value+140))
#row141 = Data_Row(int(value+141))
#row142 = Data_Row(int(value+142))
#row143 = Data_Row(int(value+143))
#row144 = Data_Row(int(value+144))
#row145 = Data_Row(int(value+145))
#row146 = Data_Row(int(value+146))
#row147 = Data_Row(int(value+147))
#row148 = Data_Row(int(value+148))
#row149 = Data_Row(int(value+149))
#row150 = Data_Row(int(value+150))
#row151 = Data_Row(int(value+151))
#row152 = Data_Row(int(value+152))
#row153 = Data_Row(int(value+153))
#row154 = Data_Row(int(value+154))
#row155 = Data_Row(int(value+155))
#row156 = Data_Row(int(value+156))
#row157 = Data_Row(int(value+157))
#row158 = Data_Row(int(value+158))
#row159 = Data_Row(int(value+159))
#row160 = Data_Row(int(value+160))
#row161 = Data_Row(int(value+161))
#row162 = Data_Row(int(value+162))
#row163 = Data_Row(int(value+163))
#row164 = Data_Row(int(value+164))
#row165 = Data_Row(int(value+165))
#row166 = Data_Row(int(value+166))
#row167 = Data_Row(int(value+167))
#row168 = Data_Row(int(value+168))
#row169 = Data_Row(int(value+169))
#row170 = Data_Row(int(value+170))
#row171 = Data_Row(int(value+171))
#row172 = Data_Row(int(value+172))
#row173 = Data_Row(int(value+173))
#row174 = Data_Row(int(value+174))
#row175 = Data_Row(int(value+175))
#row176 = Data_Row(int(value+176))
#row177 = Data_Row(int(value+177))
#row178 = Data_Row(int(value+178))
#row179 = Data_Row(int(value+179))
#row180 = Data_Row(int(value+180))
#row181 = Data_Row(int(value+181))
#row182 = Data_Row(int(value+182))
# row183 = Data_Row(int(value+183))
# row184 = Data_Row(int(value+184))
# row185 = Data_Row(int(value+185))

workbook_obj.close()
print("DEBUG: The workbook has closed")
print(row1.row_formatted)
# Launch chrome's web driver 
#driver = webdriver.Chrome()
# Bring up ascom form
#driver.get("https://www.dashboard.meraki.com")

#####################################################################################
# Login                          								            	    #
#####################################################################################
#email_address = driver.find_element_by_name("txtReqFirstName")
#email_address.send_keys(input("Enter Email:"))

#password = driver.find_element_by_name("txtReqLastName")
#password.send_keys(input("Enter Password:"))


#driver.quit()


